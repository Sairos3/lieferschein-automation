import sqlite3
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.compare import compute_stock_overview, compute_invoice_statuses


def red_outer_border() -> Border:
    red = Side(style="medium", color="FF0000")
    return Border(left=red, right=red, top=red, bottom=red)


def apply_row_outer_border(ws, row_idx: int, from_col: int, to_col: int, border: Border):
    for c in range(from_col, to_col + 1):
        ws.cell(row=row_idx, column=c).border = border


def auto_adjust_column_width(ws, min_width=10, extra_padding=3, filter_padding=4):
    """
    filter_padding: extra space for Excel filter dropdown arrow in header
    """
    for col_cells in ws.columns:
        column_letter = col_cells[0].column_letter

        header_val = ws[f"{column_letter}1"].value
        header_len = len(str(header_val)) if header_val is not None else 0

        max_len = header_len
        for cell in col_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))

        width = max_len + extra_padding + filter_padding
        ws.column_dimensions[column_letter].width = max(min_width, width)


def style_header(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True)
        # left header avoids filter overlap
        cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def apply_column_alignment(ws, align_map: dict[int, str]):
    for row in ws.iter_rows(min_row=2):
        for col_idx, align in align_map.items():
            cell = row[col_idx - 1]
            cell.alignment = Alignment(horizontal=align, vertical="center")


def export_to_excel(db_path: Path, output_path: Path) -> None:
    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row

    wb = Workbook()
    border_open = red_outer_border()

    # -----------------------
    # Delivery Notes (no prices)
    # -----------------------
    ws_dn = wb.active
    ws_dn.title = "DeliveryNotes"
    ws_dn.append(["Supplier", "TaxNo", "DeliveryNoteNo", "DeliveryDate", "CustomerNo", "OrderNo", "SourceFile"])
    style_header(ws_dn)

    dn_rows = con.execute("""
        SELECT s.name, s.tax_no, d.delivery_note_no, d.delivery_date, d.customer_no, d.order_no, d.source_pdf
        FROM delivery_notes d
        JOIN suppliers s ON s.id = d.supplier_id
        ORDER BY d.id
    """).fetchall()

    for r in dn_rows:
        ws_dn.append([r["name"], r["tax_no"], r["delivery_note_no"], r["delivery_date"], r["customer_no"], r["order_no"], r["source_pdf"]])

    apply_column_alignment(ws_dn, {1: "left", 2: "left", 3: "left", 4: "center", 5: "left", 6: "left", 7: "left"})

    # -----------------------
    # Delivery Items (no prices)
    # -----------------------
    ws_di = wb.create_sheet("DeliveryItems")
    ws_di.append(["Supplier", "TaxNo", "DeliveryNoteNo", "LineNo", "ArticleNo", "Description", "QtyDelivered"])
    style_header(ws_di)

    di_rows = con.execute("""
        SELECT s.name, s.tax_no, d.delivery_note_no, i.line_no, i.article_no, i.description, i.quantity
        FROM delivery_note_items i
        JOIN delivery_notes d ON d.id = i.delivery_note_id
        JOIN suppliers s ON s.id = d.supplier_id
        ORDER BY d.id, i.line_no
    """).fetchall()

    for r in di_rows:
        ws_di.append([r["name"], r["tax_no"], r["delivery_note_no"], r["line_no"], r["article_no"], r["description"], r["quantity"]])

    apply_column_alignment(ws_di, {1: "left", 2: "left", 3: "left", 4: "right", 5: "left", 6: "left", 7: "right"})

    # -----------------------
    # Invoices
    # -----------------------
    ws_inv = wb.create_sheet("Invoices")
    ws_inv.append(["Supplier", "TaxNo", "InvoiceNo", "InvoiceDate", "OrderNo", "CustomerNo", "Status", "SourceFile"])
    style_header(ws_inv)

    inv_rows = con.execute("""
        SELECT s.name, s.tax_no, i.invoice_no, i.invoice_date, i.order_no, i.customer_no, i.source_pdf, i.id
        FROM invoices i
        JOIN suppliers s ON s.id = i.supplier_id
        ORDER BY i.id
    """).fetchall()

    inv_status = compute_invoice_statuses(con)

    for r in inv_rows:
        invoice_id = int(r["id"])
        status = inv_status.get(invoice_id, "COMPLETE")
        ws_inv.append([r["name"], r["tax_no"], r["invoice_no"], r["invoice_date"], r["order_no"], r["customer_no"], status, r["source_pdf"]])

    apply_column_alignment(ws_inv, {1: "left", 2: "left", 3: "left", 4: "center", 5: "left", 6: "left", 7: "center", 8: "left"})

    # mark OPEN invoices with red border
    for row_idx in range(2, ws_inv.max_row + 1):
        if ws_inv.cell(row=row_idx, column=7).value == "OPEN":
            apply_row_outer_border(ws_inv, row_idx, 1, 8, border_open)

    # -----------------------
    # Invoice Items
    # -----------------------
    ws_ii = wb.create_sheet("InvoiceItems")
    ws_ii.append(["InvoiceNo", "LineNo", "ArticleNo", "Description", "QtyExpected"])
    style_header(ws_ii)

    ii_rows = con.execute("""
        SELECT i.invoice_no, ii.line_no, ii.article_no, ii.description, ii.qty_expected
        FROM invoice_items ii
        JOIN invoices i ON i.id = ii.invoice_id
        ORDER BY i.id, ii.line_no
    """).fetchall()

    for r in ii_rows:
        ws_ii.append([r["invoice_no"], r["line_no"], r["article_no"], r["description"], r["qty_expected"]])

    apply_column_alignment(ws_ii, {1: "left", 2: "right", 3: "left", 4: "left", 5: "right"})

    # -----------------------
    # Stock (Expected vs Delivered vs Open)
    # -----------------------
    ws_stock = wb.create_sheet("Stock")
    ws_stock.append(["Supplier", "TaxNo", "OrderNo", "ArticleNo", "Description", "Expected", "Delivered", "Open", "Status"])
    style_header(ws_stock)

    stock = compute_stock_overview(con)
    for r in stock:
        ws_stock.append([
            r["supplier"], r["supplier_tax_no"], r["order_no"],
            r["article_no"], r["description"],
            r["expected"], r["delivered"], r["open"], r["status"]
        ])

    apply_column_alignment(ws_stock, {1: "left", 2: "left", 3: "left", 4: "left", 5: "left", 6: "right", 7: "right", 8: "right", 9: "center"})

    # mark OPEN stock rows with red border
    for row_idx in range(2, ws_stock.max_row + 1):
        if ws_stock.cell(row=row_idx, column=9).value == "OPEN":
            apply_row_outer_border(ws_stock, row_idx, 1, 9, border_open)

    # -----------------------
    # Widths
    # -----------------------
    for ws in [ws_dn, ws_di, ws_inv, ws_ii, ws_stock]:
        auto_adjust_column_width(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print("Excel exported to:", output_path.resolve())